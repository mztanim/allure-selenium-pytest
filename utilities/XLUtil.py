from openpyxl import load_workbook

path, workbook, sheet = "", "", ""


def open_wb(filename, sheet_name):
    global path, workbook, sheet
    path = filename
    workbook = load_workbook(path)
    sheet = workbook[sheet_name]


def get_max_rows():
    return sheet.max_row


def get_max_cols():
    return sheet.max_column


def get_cell_data(row, col):
    value = sheet.cell(row=row, column=col).value
    if value is None:
        return ""
    return value


def set_cell_data(row, col, data):
    sheet.cell(row=row, column=col).value = data
    workbook.save(path)


def close_wb():
        workbook.close()

def get_col_num(header):
    for i in range(1,get_max_cols()):
        if sheet.cell(row=1, column=i).value != header:
            continue

        return i

